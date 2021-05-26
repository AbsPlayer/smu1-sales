# coding: utf-8
import requests
import bs4
from auth import login, password
from fake_useragent import UserAgent
import openpyxl
from operator import itemgetter
import time


def save_to_xlsx(filename, data):

    filename += '.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        'Номер',
        'Подъезд',
        'Кол - во комнат',
        'Cтудия',
        'Этаж',
        'Статус',
        'Площадь общая',
        'Площадь жилая',
        'Площадь кухни',
        'Отделка',
        'Кол - во раздельных санузлов',
        'Кол - во совмещенных санузлов',
        'Кол - во балконов',
        'Кол - во лоджий',
        'Цена за квартиру',
        'Цена за м2',
        'Дом'
    ]
    for index, header in enumerate(headers):
        ws.cell(row=1, column=index + 1).value = header

    # data = sorted(data, key=lambda item: item['nokv'])
    data = sorted(data, key=itemgetter('dom', 'nopodezd', 'nokv'))
    r = 2
    for item in data:
        ws.cell(row=r, column=1).value = item['nokv']
        ws.cell(row=r, column=2).value = item['nopodezd']
        ws.cell(row=r, column=3).value = item['kkomnat']
        ws.cell(row=r, column=4).value = item['studia']
        ws.cell(row=r, column=5).value = item['etag']
        ws.cell(row=r, column=6).value = item['status']
        ws.cell(row=r, column=7).value = item['sq_total']
        ws.cell(row=r, column=8).value = item['sq_jil']
        ws.cell(row=r, column=9).value = item['sq_kuh']
        ws.cell(row=r, column=10).value = item['otdelka']
        ws.cell(row=r, column=11).value = item['qtyrsu']
        ws.cell(row=r, column=12).value = item['qtyssu']
        ws.cell(row=r, column=13).value = item['qty_balkon']
        ws.cell(row=r, column=14).value = item['qty_lodgia']
        ws.cell(row=r, column=15).value = item['cenakv']
        ws.cell(row=r, column=16).value = item['cenam2']
        ws.cell(row=r, column=17).value = item['dom']
        r += 1
    wb.save(filename)

    return


start_time = time.time()
url = 'https://smu1.sales-platform.ru/realtor.php'
url_login = url + '?a=login'
url2 = 'https://smu1.sales-platform.ru/index.php?modus=friel'
ua = UserAgent().random
headers = {
    'User-Agent': ua,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Referer': url,
    'Accept-Encoding': 'gzip, deflate, br'
}
data = {"login": login, "password": password}
session = requests.Session()
session.post(url_login, data=data, headers=headers)
phpsessid = session.cookies['PHPSESSID']
loginhash = session.cookies['loginhash']
headers['Cookie'] = f'PHPSESSID={phpsessid}; login={login}; loginhash={loginhash}'
r = session.get(url2, headers=headers)
if r.status_code != requests.codes.ok:
    print("Сайт не отвечает после залогинивания")
    quit()
soup = bs4.BeautifulSoup(r.text, 'html.parser')
gks = [i['value'] for i in soup.find(id="gk").find_all('option')]
lengks = len(gks)
for i in range(lengks):
    print(f'Обрабатывается {i+1}-й жилой комплекс из {lengks}')
    data = []
    gk = gks[i]
    url_gk = f'https://smu1.sales-platform.ru/index.php?gk={gk}&modus=friel'
    r = requests.get(url_gk, headers=headers)
    if r.status_code != requests.codes.ok:
        print("Сайт не отвечает при парсинге ЖК")
        quit()
    soup = bs4.BeautifulSoup(r.text, 'html.parser')
    buildings = [i['value'] for i in soup.find(attrs={'name': 'house'}).find_all('option') \
                 if 'парковка' not in i.text.lower()]
    for building in buildings:
        url_building = f'https://smu1.sales-platform.ru/index.php?a=shah&gk={gk}&modus=friel&idhouse={building}'
        headers['Cookie'] = f'PHPSESSID={phpsessid}; login={login}; loginhash={loginhash}; sidebarstate=min'
        r = requests.get(url_building, headers=headers)
        if r.status_code != requests.codes.ok:
            print("Сайт не отвечает при парсинге квартир")
            quit()
        soup = bs4.BeautifulSoup(r.text, 'html.parser')
        podezds = soup.find_all(id='modedivsh')
        for podezd in podezds:
            nopodezd = podezd.h3.text
            kvartiras = podezd.find_all('button', id='plat')
            for kvartira in kvartiras:
                try:
                    info = kvartira.attrs['onclick'][21:-1].split("',")
                    nokv = int(kvartira.attrs['data-apartment'])
                    etag = int(info[12].strip().strip("'").split("/")[0])

                    try:
                        sq_total = float(info[1].strip().strip("'"))
                    except:
                        sq_total = info[1].strip().strip("'")

                    try:
                        sq_jil = float(info[2].strip().strip("'"))
                    except:
                        sq_jil = info[2].strip().strip("'")

                    try:
                        sq_kuh = float(info[3].strip().strip("'"))
                    except:
                        sq_kuh = info[3].strip().strip("'")

                    try:
                        kkomnat = int(info[7].strip().strip("'"))
                    except:
                        kkomnat = info[7].strip().strip("'")

                    try:
                        cenakv = float(''.join(info[4].strip().strip("'").split()))
                    except:
                        cenakv = ''.join(info[4].strip().strip("'").split())

                    try:
                        cenam2 = float(''.join(info[6].strip().strip("'").split()))
                    except:
                        cenam2 = ''.join(info[6].strip().strip("'").split())

                    dom = info[16].strip().strip("'")
                    try:
                        sq_balkon = float(info[15].split(":")[1].strip()[:-5])
                    except:
                        sq_balkon = ''
                    if info[14].strip().strip("'").lower() == 'свободна':
                        status = 'Свободно'
                    elif info[14].strip().strip("'").lower() == 'забронирована':
                        status = 'Забронировано'
                    else:
                        status = 'Продано'
                    # studia = 'Да' if 'СТУДИЯ' in info[19] else ''
                    studia = 'Да' if info[13].strip().strip("'").lower() == 'да' else ''
                    otdelka = ''
                    if info[30].strip().strip("'") != '0':
                        # otdelka = soup.find(id='priceodin').next.next.next
                        otdelka = 'черновая'
                    if info[26].strip().strip("'") != '0':
                        otdelka = soup.find(id='pricedva').next.next.next
                    if info[28].strip().strip("'") != '0':
                        otdelka = soup.find(id='pricetri').next.next.next
                    if info[36].strip().strip("'") != '0':
                        otdelka = soup.find(id='pricechet').next.next.next

                    data.append({'nokv': nokv,
                                 'nopodezd': nopodezd,
                                 'kkomnat': kkomnat,
                                 'studia': studia,
                                 'etag': etag,
                                 'status': status,
                                 'sq_total': sq_total,
                                 'sq_jil': sq_jil,
                                 'sq_kuh': sq_kuh,
                                 'otdelka': otdelka,
                                 'qtyrsu': '',
                                 'qtyssu': '',
                                 'qty_balkon': '',
                                 'qty_lodgia': '',
                                 'cenakv': cenakv,
                                 'cenam2': cenam2,
                                 'dom': dom
                                 })
                except:
                    pass

    save_to_xlsx(gk, data)
t = time.time() - start_time
# print(f"{t} seconds")
print('All done!!!')
